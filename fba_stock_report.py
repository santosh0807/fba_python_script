import pandas as pd
import math
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# -------------------------------------------------
# CONFIGURATION (unchanged)
# -------------------------------------------------
google_sheet_id = "1KDQJM6ZIVZCkFoylfoXqsbead93zwjzCc61wHGf2GSM"
main_gid = "712927586"
orders_gid = "0"

stock_csv_url = f"https://docs.google.com/spreadsheets/d/{google_sheet_id}/export?format=csv&gid={main_gid}"
orders_csv_url = f"https://docs.google.com/spreadsheets/d/{google_sheet_id}/export?format=csv&gid={orders_gid}"

output_file = "/home/santosh/Documents/Python/Scripts/FBA/FBA Report/final_output.xlsx"

# -------------------------------------------------
# READ MAIN STOCK SHEET
# -------------------------------------------------
print("Loading main stock data (with stock + sales)...")
df_stock = pd.read_csv(stock_csv_url, low_memory=False)
df_stock["Receipt Date"] = pd.to_datetime(df_stock["Receipt Date"], errors="coerce")

# -------------------------------------------------
# FC LIST
# -------------------------------------------------
fc_order = ["RNR0","BOM5","BOM7","PNQ3","SBLY","BLR8","BLR7","BLR5","HYD8","HYD3",
            "CJB1","MAA4","LKO1","DEL5","DEL4","DEL2","DED4","CCX2","CCX1","GAX1","DEX3","IDX2","NAX1","CCX4","ZSHA","FMAB","FHYE","DED3","DED5","JPX2"]

# -------------------------------------------------
# SHEET 1 — Replenishment_Needed (NOW WITH Receipt Date)
# -------------------------------------------------
print("Generating Replenishment_Needed & Low_Selling_Items...")
sheet1 = []
for _, row in df_stock.iterrows():
    total_stock = row["Total Stock (All FCs)"]
    total_sale = row["Total Sale (All FCs)"]
    if pd.isna(total_sale) or total_sale <= 0: continue
    required_45 = math.ceil((total_sale / 30) * 45)
    if total_stock < required_45:
        needed = required_45 - total_stock
        fc_vals = [int(row[col]) if col in row and pd.notna(row[col]) else 0 for col in [f"{fc} Stock" for fc in fc_order]]
        rec_date = row["Receipt Date"].date() if pd.notna(row["Receipt Date"]) else ""
        sheet1.append([row["SKU"], row["ASIN"], row["Product Name"], rec_date, total_stock, total_sale, required_45, needed] + fc_vals)

df_sheet1 = pd.DataFrame(sheet1, columns=[
    "SKU","ASIN","Product Name","Receipt Date","Current Stock","Total Sale","Required (45 Days)","Replenishment Needed"
] + [f"{fc} Stock" for fc in fc_order])

# -------------------------------------------------
# SHEET 2 — Low_Selling_Items (unchanged)
# -------------------------------------------------
sheet2 = []
for _, row in df_stock.iterrows():
    stock = row["Total Stock (All FCs)"]
    sale = row["Total Sale (All FCs)"]
    if pd.isna(stock) or stock <= 0: continue
    sold_pct = (sale / stock) * 100
    if sold_pct < 30:
        fc_vals = [int(row[col]) if col in row and pd.notna(row[col]) else 0 for col in [f"{fc} Stock" for fc in fc_order]]
        rec_date = row["Receipt Date"].date() if pd.notna(row["Receipt Date"]) else ""
        sheet2.append([row["SKU"], row["ASIN"], row["Product Name"], rec_date, stock, sale, round(sold_pct,2)] + fc_vals)

df_sheet2 = pd.DataFrame(sheet2, columns=["SKU","ASIN","Product Name","Receipt Date","Stock","Sale","% Sold"] + [f"{fc} Stock" for fc in fc_order])
df_sheet2 = df_sheet2.sort_values("Receipt Date")

# -------------------------------------------------
# SHEET 3 — QWQC (unchanged)
# -------------------------------------------------
print("Generating QWQC...")
df_orders = pd.read_csv(orders_csv_url, low_memory=False)
df_orders['Order_Date'] = pd.to_datetime(df_orders.iloc[:, 0].astype(str).str.strip(), format='%Y-%m-%d', errors='coerce')
df_orders = df_orders.dropna(subset=['Order_Date']).copy()
df_orders['SKU'] = df_orders.iloc[:, 1].astype(str).str.strip()
df_orders['FC_Original'] = df_orders.iloc[:, 4].astype(str).str.strip().str.upper()
df_orders['Qty'] = pd.to_numeric(df_orders.iloc[:, 5], errors='coerce').fillna(0).astype(int)
df_orders = df_orders[(df_orders['SKU'].str.len() > 4) & (df_orders['Qty'] > 0)]

cutoff_30 = datetime.now().date() - timedelta(days=30)
df_qwqc = df_orders[
    (df_orders['Order_Date'].dt.date >= cutoff_30) &
    (df_orders['FC_Original'].str.contains('QWQC', case=False, na=False))
]
qwqc_summary = df_qwqc.groupby('SKU')['Qty'].sum().reset_index().rename(columns={'SKU': 'QWQC'})
qwqc_summary = qwqc_summary.sort_values('Qty', ascending=False)
qwqc_summary["BOM5 + BOM7 Stock"] = 0
for idx, row in qwqc_summary.iterrows():
    match = df_stock[df_stock["SKU"] == row["QWQC"]]
    if not match.empty:
        bom5 = int(match.iloc[0].get("BOM5 Stock", 0) or 0)
        bom7 = int(match.iloc[0].get("BOM7 Stock", 0) or 0)
        qwqc_summary.at[idx, "BOM5 + BOM7 Stock"] = bom5 + bom7
qwqc_summary.loc[len(qwqc_summary)] = ["GRAND TOTAL", qwqc_summary["Qty"].sum(), qwqc_summary["BOM5 + BOM7 Stock"].sum()]

# -------------------------------------------------
# SHEET 4 — ZONEWISE (NOW WITH Receipt Date)
# -------------------------------------------------
print("Generating zonewise sheet (with Receipt Date)...")

df_zonewise = df_stock[['SKU', 'ASIN', 'Product Name',
                        'Total Stock (All FCs)', 'Total Sale (All FCs)', 'Receipt Date']].copy()
df_zonewise.rename(columns={'Total Stock (All FCs)': 'Current Stock',
                            'Total Sale (All FCs)': 'Total Sale'}, inplace=True)
df_zonewise['Receipt Date'] = df_zonewise['Receipt Date'].dt.date

for fc in fc_order:
    stock_col = f"{fc} Stock"
    sale_col = f"{fc} Sale"
    
    df_zonewise[stock_col] = df_stock.get(stock_col, 0).fillna(0).astype(int)
    df_zonewise[sale_col] = df_stock.get(sale_col, 0).fillna(0).astype(int)
    
    rep_col = "45 day replenish" if fc == "RNR0" else f"{fc} 45 day replenish"
    df_zonewise[rep_col] = (df_zonewise[sale_col] / 30 * 45).apply(math.ceil).astype(int)

# Final column order — Receipt Date added right after Product Name
final_cols = ['SKU', 'ASIN', 'Product Name', 'Receipt Date', 'Current Stock', 'Total Sale']
for fc in fc_order:
    final_cols += [f"{fc} Stock", f"{fc} Sale", "45 day replenish" if fc=="RNR0" else f"{fc} 45 day replenish"]
df_zonewise = df_zonewise[final_cols]

# -------------------------------------------------
# SAVE & FORMAT (unchanged + auto column width will handle new column)
# -------------------------------------------------
print("Saving final report...")
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    df_sheet1.to_excel(writer, "Replenishment_Needed", index=False)
    df_sheet2.to_excel(writer, "Low_Selling_Items", index=False)
    qwqc_summary.to_excel(writer, "QWQC", index=False)
    df_zonewise.to_excel(writer, "zonewise", index=False)

wb = load_workbook(output_file)
yellow_fill = PatternFill(start_color="FFF59D", fill_type="solid")
border = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

for ws in wb:
    for cell in ws[1]:
        cell.fill = yellow_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = max_len + 4

# Highlight 45-day replenish columns (unchanged)
ws = wb["zonewise"]
for i, cell in enumerate(ws[1], 1):
    if "45 day replenish" in str(cell.value):
        for r in ws.iter_rows(min_row=2, min_col=i, max_col=i):
            if r[0].value and r[0].value > 0:
                r[0].fill = PatternFill(start_color="FFFF99", fill_type="solid")

wb.save(output_file)

print("\nFINAL REPORT WITH RECEIPT DATE ADDED!")
print("→ Replenishment_Needed: Receipt Date added")
print("→ zonewise: Receipt Date added (after Product Name)")
print("Everything else 100% unchanged & perfect")
print(f"File saved: {output_file}")